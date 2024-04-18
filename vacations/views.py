from typing import Any
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import VacationRequestForm, VacationRequestUpdateForm
from lk.models import User
from vacations.models import VacationRequest
from django.core.exceptions import ValidationError
from django.views.generic.edit import FormView
from django.views.generic import ListView, View, TemplateView, UpdateView
from django.urls import reverse_lazy
import datetime
from .const import *
from django.conf import settings
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.shortcuts import render
from openpyxl import load_workbook

def import_from_excel(request):
    if request.method == 'POST' and request.FILES:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_id, date, days = row
            user = get_user_by_id(user_id)
            start_date = date
            finish_date = start_date + datetime.timedelta(days=days)
            if user:
                VacationRequest.objects.create(user=user, start_date=start_date, finish_date= finish_date, days=days)
                if days >=14:
                    user.is_14 = True
                user.available_days -= days
                user.save() 

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')


def get_user_by_id(id):
    try:
        user = User.objects.get(pk=id)
        return user
    except User.DoesNotExist:
        return None

class VacationRequestView(TemplateView, FormView):
    http_method_names = ['get', 'post']
    template_name = 'vacation_request.html'
    form_class = VacationRequestForm 
    success_url = reverse_lazy('lk')
    vacation_days = None
    
    def get(self, request, *args, **kwargs):
        user = get_user_by_id(request.user.pk)
        context = self.get_context_data(**kwargs)
        form = self.get_form(self.form_class)
        context.update({
            'form': form,
            'user': user
        })        
        return render(request, 'vacation_request.html', context)

    def form_validation(self, form, request):
        if form.is_valid():
            started_date = datetime.date.fromisoformat(form['start_date'].data)
            finish_date = datetime.date.fromisoformat(form['finish_date'].data)
            if started_date > finish_date:
                messages.error(request, ValidationError("Первый день отпуска должен быть раньше последнего").message)
                return False

            vacation_days = (finish_date - started_date).days + 1
            if request.user.available_days >= vacation_days:
                if request.user.is_14 or vacation_days >=14 or request.user.available_days - vacation_days >=14:
                    self.vacation_days = vacation_days
                    return True
                else:
                    messages.error(request, ValidationError("Должен быть 14 дневный отпуск").message)
                    return False
            else:
                messages.error(request, ValidationError("Не хватает дней").message)
                return False
        return False

        
    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        form = self.get_form(self.form_class)        
        if self.form_validation(form, request):
            try:
                obj = form.save(user=request.user, vacation_days=self.vacation_days)
                user = get_user_by_id(request.user.pk)
                user.available_days = user.available_days - self.vacation_days
                if self.vacation_days >= 14:
                    user.is_14=True
                user.save() 
                msg = EmailMessage('Запрос на отпуск', f'{obj}', to=[f'{user.user_boss.email}'])
                msg.send()              
                messages.success(request, "Запрос отправлен")
                return self.form_valid(form)
            except:
                return self.form_invalid(form)
        else:
            return self.form_invalid(form)


class VacationRequestUpdateView(UpdateView):
    model = VacationRequest
    template_name = 'vacation_approval.html'
    fields = ['status', 'comment']
    success_url = reverse_lazy('lk')

    def post(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
       
        if request.POST['status'] == REJECTED:
            vac_req = VacationRequest.objects.get(pk=kwargs['pk'])
            user = get_user_by_id(vac_req.user.pk)
            user.available_days += vac_req.days
            user.save()

        return super().post(request, *args, **kwargs)
        
